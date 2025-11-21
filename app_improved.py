# app_improved.py
# Sistema de RelatÃ³rios & Planejamento de ProduÃ§Ã£o
# VersÃ£o unificada usando apenas template_estoque (Google Sheets - SOMENTE LEITURA)

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import numpy as np

# ==============================================================================
# CONFIGURAÃ‡Ã•ES GERAIS
# ==============================================================================

# ðŸ”— CONFIG DO GOOGLE SHEETS (APENAS LEITURA)
# Se precisar mudar, sÃ³ troque o ID ou o nome da aba.
GOOGLE_SHEET_ID = "1PpiMQingHf4llA03BiPIuPJPIZqul4grRU_emWDEK1o"
TEMPLATE_SHEET_NAME = "template_estoque"  # nome da aba onde estÃ¡ o estoque

st.set_page_config(
    page_title="Pure & Posh Baby - Vendas â†’ Estoque â†’ ProduÃ§Ã£o",
    page_icon="ðŸ‘‘",
    layout="wide"
)

st.markdown(
    """
    <style>
    .centered-title { text-align: center; width: 100%; margin: 0 auto; }
    @media (max-width: 768px) { .centered-title { text-align: center; } }
    </style>
    """,
    unsafe_allow_html=True
)

st.markdown('<div class="centered-title">', unsafe_allow_html=True)
st.title("ðŸ‘‘ Sistema de RelatÃ³rios & Planejamento de ProduÃ§Ã£o")
st.markdown("**Pure & Posh Baby â€” Vendas â†’ Estoque â†’ ProduÃ§Ã£o**")
st.markdown('</div>', unsafe_allow_html=True)

# ==============================================================================
# ESTADO
# ==============================================================================

if "df_estoque" not in st.session_state:
    st.session_state["df_estoque"] = None

if "template_carregado" not in st.session_state:
    st.session_state["template_carregado"] = False

# ==============================================================================
# FUNÃ‡Ã•ES AUXILIARES
# ==============================================================================

@st.cache_data
def load_template_from_google(sheet_id: str, sheet_name: str):
    """
    LÃª o template_estoque direto do Google Sheets em modo SOMENTE LEITURA.

    Importante:
    - A planilha precisa permitir leitura pÃºblica OU
      estar compartilhada de forma que o servidor do Streamlit consiga ler.
    - Este mÃ©todo apenas faz um GET no link de exportaÃ§Ã£o, nÃ£o tem permissÃ£o de escrita.
    """
    # URL padrÃ£o de exportaÃ§Ã£o em XLSX
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    df = pd.read_excel(url, sheet_name=sheet_name)
    return df

@st.cache_data
def load_excel(file, sheet_name=None):
    """Carrega um Excel em DataFrame, com cache (usado para vendas)."""
    return pd.read_excel(file, sheet_name=sheet_name)

def normalizar_colunas(df):
    df = df.copy()
    df.columns = (
        df.columns
        .str.strip()
        .str.lower()
        .str.replace(" ", "_")
        .str.replace("Ã£", "a")
        .str.replace("Ã¡", "a")
        .str.replace("Ã©", "e")
        .str.replace("Ã§", "c")
    )
    return df

def bool_from_any(x):
    if pd.isna(x):
        return False
    s = str(x).strip().lower()
    return s in ["1", "true", "sim", "yes", "y"]

def split_list(texto):
    if pd.isna(texto):
        return []
    return [t.strip() for t in str(texto).split(",") if t.strip()]

def get_categoria_ordem(semi_nome):
    """
    Define ordem dos semis pela descriÃ§Ã£o:
    1 = Manga Longa
    2 = Manga Curta Menina
    3 = Manga Curta Menino
    4 = MijÃ£o
    Depois por cor (Branco, Off, Rosa, Azul, Vermelho, Marinho, outros)
    Depois por tamanho (RN, P, M, G).
    """
    s = str(semi_nome).lower()

    # tipo
    if "manga longa" in s:
        cat = 1
    elif "manga curta" in s and ("menina" in s or "fem" in s):
        cat = 2
    elif "manga curta" in s and ("menino" in s or "masc" in s):
        cat = 3
    elif "mijao" in s or "mijÃ£o" in s:
        cat = 4
    else:
        cat = 9

    # cor
    if "branco" in s and "off" not in s:
        cor = 1
    elif "off" in s:
        cor = 2
    elif "rosa" in s:
        cor = 3
    elif "azul" in s:
        cor = 4
    elif "vermelho" in s or "verme" in s:
        cor = 5
    elif "marinho" in s:
        cor = 6
    else:
        cor = 9

    # tamanho
    if "-rn" in s or " rn" in s:
        tam = 1
    elif "-p" in s or " p" in s:
        tam = 2
    elif "-m" in s or " m" in s:
        tam = 3
    elif "-g" in s or " g" in s:
        tam = 4
    else:
        tam = 9

    return cat, cor, tam

def gerar_excel_semis_golas(relatorio_linhas):
    """
    Gera um Excel hierÃ¡rquico:
    - Linha de Semi (negrito, cor de fundo)
    - Linhas de Golas logo abaixo, com leve indentaÃ§Ã£o
    """
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Produzir Hoje"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    semi_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Item", "Qtd NecessÃ¡ria", "Estoque Atual", "Falta"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row = 2
    for linha in relatorio_linhas:
        is_semi = linha["tipo"] == "semi"
        for col_idx, key in enumerate(["item", "qtd_necessaria", "estoque_atual", "falta"], 1):
            cell = ws.cell(row=row, column=col_idx, value=linha.get(key, ""))
            cell.border = border
            if is_semi:
                if col_idx == 1:
                    cell.font = Font(bold=True)
                cell.fill = semi_fill
        row += 1

    # Ajuste de largura
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output)
    output.seek(0)
    return output

def gerar_excel_simples(df, sheet_name="Relatorio"):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output)
    output.seek(0)
    return output

# ==============================================================================
# 1. CARREGAR TEMPLATE_ESTOQUE DO GOOGLE (SOMENTE LEITURA)
# ==============================================================================

st.header("1. ConfiguraÃ§Ã£o Inicial â€” template_estoque (Google Sheets - somente leitura)")

with st.expander("ðŸ“˜ Como funciona o template_estoque", expanded=True):
    st.markdown(
        """
        **Ã‰ a planilha ÃšNICA de estoque usada por toda a operaÃ§Ã£o.**

        Este app **nÃ£o faz upload**, nÃ£o grava, nÃ£o altera cÃ©lulas.
        Ele apenas lÃª o conteÃºdo da aba configurada em `TEMPLATE_SHEET_NAME`
        dentro da planilha cujo ID estÃ¡ em `GOOGLE_SHEET_ID`.

        Colunas principais esperadas:

        - `codigo` â†’ cÃ³digo do item (produto pronto, semi, gola, bordado, kit, etc.)
        - `nome` â†’ descriÃ§Ã£o legÃ­vel
        - `categoria` â†’ ex: *Bodys Prontos*, *Semi Manga Longa*, *Golas*, *Bordados*, *Conjuntos*â€¦
        - `estoque_atual` â†’ quantidade em estoque (pode ser negativa)
        - `eh_kit` â†’ (opcional) â€œSimâ€ quando for kit de outros cÃ³digos
        - `componentes` â†’ (opcional) lista de cÃ³digos dos componentes do kit, separados por vÃ­rgula
        - `quantidades` â†’ (opcional) lista de quantidades correspondentes aos componentes (mesma ordem)

        Para **produÃ§Ã£o**, Ã© recomendado ter:

        - `semi_codigo` â†’ cÃ³digo do semi usado em cada produto pronto
        - `gola_codigo` â†’ cÃ³digo da gola pronta usada em cada produto
        - `bordado_codigo` â†’ cÃ³digo do bordado (quando a gola depender de bordado)

        ðŸ”’ **Importante**: este app sÃ³ **LÃŠ** o template_estoque.  
        Quem continua comandando o estoque Ã© o seu outro app.
        """
    )

col_a, col_b = st.columns([1, 3])
with col_a:
    if st.button("ðŸ”„ Recarregar do Google Sheets"):
        load_template_from_google.clear()
        st.session_state["template_carregado"] = False
        st.rerun()

if not st.session_state["template_carregado"]:
    try:
        df_est = load_template_from_google(GOOGLE_SHEET_ID, TEMPLATE_SHEET_NAME)
        df_est = normalizar_colunas(df_est)

        colunas_obrigatorias = ["codigo", "nome", "categoria", "estoque_atual"]
        faltando = [c for c in colunas_obrigatorias if c not in df_est.columns]

        if faltando:
            st.error(
                "âŒ O template_estoque (aba "
                f"`{TEMPLATE_SHEET_NAME}`) precisa ter as colunas: {', '.join(colunas_obrigatorias)}. "
                f"Faltando: {', '.join(faltando)}"
            )
        else:
            # Garante colunas opcionais
            for col in ["eh_kit", "componentes", "quantidades",
                        "semi_codigo", "gola_codigo", "bordado_codigo"]:
                if col not in df_est.columns:
                    df_est[col] = ""

            st.session_state["df_estoque"] = df_est
            st.session_state["template_carregado"] = True

            total_itens = len(df_est)
            total_kits = df_est["eh_kit"].apply(bool_from_any).sum()
            total_mapeados = df_est["semi_codigo"].astype(str).str.strip().ne("").sum()

            st.success(
                f"âœ… template_estoque lido do Google Sheets com **{total_itens} itens**, "
                f"**{total_kits} kits** e **{total_mapeados} produtos** mapeados em semi/gola/bordado."
            )

            st.dataframe(df_est.head(20))
    except Exception as e:
        st.error(f"Erro ao ler template_estoque do Google Sheets: {e}")

# ==============================================================================
# 2. PROCESSAR VENDAS DO DIA
# ==============================================================================

if not st.session_state["template_carregado"]:
    st.info("âž¡ Antes, garanta que o template_estoque foi carregado com sucesso.")
else:
    st.header("2. Processar Vendas do Dia")

    with st.expander("ðŸ“‘ Formato da planilha de vendas", expanded=True):
        st.markdown(
            """
            Esperado um arquivo **Excel (.xlsx)** com pelo menos:

            - Coluna `CÃ³digo` ou `codigo`
            - Coluna `Quantidade` ou `quantidade`

            ðŸ‘‰ VocÃª pode exportar a planilha diÃ¡ria do Mercado Livre / Shopee / etc.  
            O app vai **somar por cÃ³digo** e usar apenas os cÃ³digos que realmente venderam.
            """
        )

    uploaded_vendas = st.file_uploader(
        "ðŸ“‚ Envie a planilha de vendas do dia",
        type=["xlsx"],
        key="vendas_file",
    )

    if uploaded_vendas:
        try:
            df_vendas = load_excel(uploaded_vendas)
            df_vendas = normalizar_colunas(df_vendas)

            # Descobrir colunas de cÃ³digo e quantidade
            col_codigo = None
            for c in ["codigo", "cÃ³digo", "cod"]:
                if c in df_vendas.columns:
                    col_codigo = c
                    break

            col_qtd = None
            for c in ["quantidade", "qtd", "qtde"]:
                if c in df_vendas.columns:
                    col_qtd = c
                    break

            if not col_codigo or not col_qtd:
                st.error(
                    "âŒ A planilha de vendas precisa ter uma coluna de **cÃ³digo** "
                    "(`codigo`, `cÃ³digo` ou `cod`) e uma de **quantidade** "
                    "(`quantidade`, `qtd` ou `qtde`)."
                )
            else:
                df_vendas = df_vendas[[col_codigo, col_qtd]].rename(
                    columns={col_codigo: "codigo", col_qtd: "quantidade"}
                )
                df_vendas["quantidade"] = pd.to_numeric(df_vendas["quantidade"], errors="coerce").fillna(0)
                df_vendas = df_vendas.groupby("codigo", as_index=False)["quantidade"].sum()
                df_vendas = df_vendas[df_vendas["quantidade"] > 0]

                st.subheader("ðŸ“Š Vendas consolidadas por cÃ³digo")
                st.dataframe(df_vendas)

                # --------------------------------------------------------------
                # 2.1. SITUAÃ‡ÃƒO DO PRODUTO PRONTO (FALTA PARA PRODUÃ‡ÃƒO)
                # --------------------------------------------------------------

                df_est = st.session_state["df_estoque"].copy()
                est_map = df_est.set_index("codigo")["estoque_atual"].to_dict()
                nome_map = df_est.set_index("codigo")["nome"].to_dict()

                df_vendas["nome"] = df_vendas["codigo"].map(nome_map).fillna("âš  CÃ³digo nÃ£o cadastrado")
                df_vendas["estoque_atual"] = df_vendas["codigo"].map(est_map).fillna(0)
                df_vendas["falta_produto"] = (df_vendas["quantidade"] - df_vendas["estoque_atual"]).clip(lower=0)

                df_produtos_faltantes = df_vendas[df_vendas["falta_produto"] > 0].copy()

                st.subheader("ðŸ“¦ SituaÃ§Ã£o de Produtos Prontos (somente faltantes)")
                if df_produtos_faltantes.empty:
                    st.success("âœ… NÃ£o hÃ¡ falta de produto pronto para os cÃ³digos desta venda.")
                else:
                    st.dataframe(df_produtos_faltantes[["codigo", "nome", "quantidade",
                                                        "estoque_atual", "falta_produto"]])

                    excel_produtos = gerar_excel_simples(
                        df_produtos_faltantes[["codigo", "nome", "quantidade",
                                               "estoque_atual", "falta_produto"]],
                        sheet_name="Produtos_Prontos"
                    )
                    st.download_button(
                        "ðŸ’¾ Baixar relatÃ³rio de Produtos Prontos (faltantes)",
                        data=excel_produtos,
                        file_name="produtos_prontos_faltantes.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                # --------------------------------------------------------------
                # 2.2. EXPLOSÃƒO EM INSUMOS (SEMI, GOLA, BORDADO)
                #     - SÃ³ para os produtos que realmente faltam
                # --------------------------------------------------------------

                st.subheader("ðŸ§© ExplosÃ£o em Insumos (apenas o que realmente precisa produzir)")

                df_est_index = df_est.set_index("codigo")

                # dicionÃ¡rios acumuladores
                semis_dict = {}       # semi_codigo -> {nome, qtd}
                golas_dict = {}       # (semi_codigo, gola_codigo) -> {nomes, qtd}
                bordados_dict = {}    # bordado_codigo -> {nome, qtd}
                erros_codigos = []

                def processar_codigo(codigo, multiplicador):
                    """Recursivamente: kit â†’ componentes â†’ produto simples â†’ insumos."""
                    if codigo not in df_est_index.index:
                        erros_codigos.append(codigo)
                        return

                    row = df_est_index.loc[codigo]
                    eh_kit = bool_from_any(row.get("eh_kit", ""))

                    if eh_kit:
                        componentes = split_list(row.get("componentes", ""))
                        quantidades = split_list(row.get("quantidades", ""))

                        # Se sÃ³ vier 1 quantidade, aplica para todos; senÃ£o, pareia
                        if len(quantidades) == 1 and len(componentes) > 1:
                            qs = [float(quantidades[0])] * len(componentes)
                        elif len(quantidades) == len(componentes):
                            qs = [float(q) for q in quantidades]
                        else:
                            # fallback: tudo com quantidade 1
                            qs = [1.0] * len(componentes)

                        for comp_cod, q in zip(componentes, qs):
                            if comp_cod:
                                processar_codigo(comp_cod, multiplicador * q)
                    else:
                        # Produto simples â†’ olhar semi / gola / bordado
                        semi_cod = str(row.get("semi_codigo", "")).strip()
                        gola_cod = str(row.get("gola_codigo", "")).strip()
                        bord_cod = str(row.get("bordado_codigo", "")).strip()

                        # SEMI
                        if semi_cod:
                            semi_nome = df_est_index.loc[semi_cod]["nome"] if semi_cod in df_est_index.index else semi_cod
                            if semi_cod not in semis_dict:
                                semis_dict[semi_cod] = {
                                    "semi_codigo": semi_cod,
                                    "semi_nome": semi_nome,
                                    "qtd_necessaria": 0.0,
                                }
                            semis_dict[semi_cod]["qtd_necessaria"] += multiplicador

                        # GOLA (casada com o semi se existir)
                        if gola_cod:
                            gola_nome = df_est_index.loc[gola_cod]["nome"] if gola_cod in df_est_index.index else gola_cod
                            chave_gola = (semi_cod, gola_cod)
                            if chave_gola not in golas_dict:
                                golas_dict[chave_gola] = {
                                    "semi_codigo": semi_cod,
                                    "semi_nome": semis_dict.get(semi_cod, {}).get("semi_nome", semi_cod),
                                    "gola_codigo": gola_cod,
                                    "gola_nome": gola_nome,
                                    "qtd_necessaria": 0.0,
                                }
                            golas_dict[chave_gola]["qtd_necessaria"] += multiplicador

                        # BORDADO (independente)
                        if bord_cod:
                            bord_nome = df_est_index.loc[bord_cod]["nome"] if bord_cod in df_est_index.index else bord_cod
                            if bord_cod not in bordados_dict:
                                bordados_dict[bord_cod] = {
                                    "bordado_codigo": bord_cod,
                                    "bordado_nome": bord_nome,
                                    "qtd_necessaria": 0.0,
                                }
                            bordados_dict[bord_cod]["qtd_necessaria"] += multiplicador

                # Rodar explosÃ£o sÃ³ para produtos com falta
                for _, row in df_produtos_faltantes.iterrows():
                    cod = row["codigo"]
                    falta = float(row["falta_produto"])
                    if falta > 0:
                        processar_codigo(cod, falta)

                if erros_codigos:
                    st.warning(
                        "âš  Alguns cÃ³digos das vendas nÃ£o foram encontrados no template_estoque "
                        "(ou em seus componentes / mapa de insumos):\n\n"
                        + ", ".join(sorted(set(erros_codigos)))
                    )

                # --------------------------------------------------------------
                # 2.3. RELATÃ“RIO SEMI + GOLAS CASADOS
                # --------------------------------------------------------------

                if not semis_dict:
                    st.success("âœ… Nenhum insumo de produÃ§Ã£o foi identificado (sem semi/gola/bordado).")
                else:
                    # map de estoque por codigo
                    est_map_full = df_est.set_index("codigo")["estoque_atual"].to_dict()

                    # DataFrame de semis para ordenar
                    df_semis = pd.DataFrame(semis_dict.values())
                    df_semis[["cat", "cor", "tam"]] = df_semis["semi_nome"].apply(
                        lambda x: pd.Series(get_categoria_ordem(x))
                    )
                    df_semis = df_semis.sort_values(["cat", "cor", "tam", "semi_nome"])

                    # ordenar golas dentro de cada semi
                    df_golas = pd.DataFrame(golas_dict.values()) if golas_dict else pd.DataFrame(
                        columns=["semi_codigo", "semi_nome", "gola_codigo", "gola_nome", "qtd_necessaria"]
                    )

                    relatorio_linhas = []

                    for _, srow in df_semis.iterrows():
                        semi_cod = srow["semi_codigo"]
                        semi_nome = srow["semi_nome"]
                        qtd_semis = float(srow["qtd_necessaria"])
                        estoque_semi = float(est_map_full.get(semi_cod, 0))
                        falta_semi = max(qtd_semis - estoque_semi, 0)

                        relatorio_linhas.append(
                            {
                                "tipo": "semi",
                                "item": f"Semi {semi_nome}",
                                "qtd_necessaria": qtd_semis,
                                "estoque_atual": estoque_semi,
                                "falta": falta_semi,
                            }
                        )

                        # golas casadas com este semi
                        if not df_golas.empty:
                            sub = df_golas[df_golas["semi_codigo"] == semi_cod].copy()
                            sub = sub.sort_values("gola_nome")
                            for _, grow in sub.iterrows():
                                gola_cod = grow["gola_codigo"]
                                gola_nome = grow["gola_nome"]
                                qtd_gola = float(grow["qtd_necessaria"])
                                estoque_gola = float(est_map_full.get(gola_cod, 0))
                                falta_gola = max(qtd_gola - estoque_gola, 0)

                                relatorio_linhas.append(
                                    {
                                        "tipo": "gola",
                                        "item": f"  Gola: {gola_nome}",
                                        "qtd_necessaria": qtd_gola,
                                        "estoque_atual": estoque_gola,
                                        "falta": falta_gola,
                                    }
                                )

                    # Mostrar tabela no app
                    st.subheader("ðŸ§µ Produzir Hoje â€” SEMIS casados com suas GOLAS")
                    df_relatorio_semis_golas = pd.DataFrame(relatorio_linhas)
                    st.dataframe(df_relatorio_semis_golas[["item", "qtd_necessaria",
                                                           "estoque_atual", "falta"]])

                    # Download Excel hierÃ¡rquico
                    excel_semis_golas = gerar_excel_semis_golas(relatorio_linhas)
                    st.download_button(
                        "ðŸ’¾ Baixar 'Produzir Hoje â€” Semis & Golas' (Excel)",
                        data=excel_semis_golas,
                        file_name="produzir_hoje_semis_golas.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                # --------------------------------------------------------------
                # 2.4. RELATÃ“RIO DE BORDADOS (OPCIONAL)
                # --------------------------------------------------------------
                st.subheader("ðŸŽ¨ Produzir Hoje â€” BORDADOS (quando mapeados)")

                if not bordados_dict:
                    st.info("Nenhum bordado foi mapeado (coluna `bordado_codigo`).")
                else:
                    df_bord = pd.DataFrame(bordados_dict.values())
                    df_bord["estoque_atual"] = df_bord["bordado_codigo"].map(est_map_full).fillna(0).astype(float)
                    df_bord["falta"] = (df_bord["qtd_necessaria"] - df_bord["estoque_atual"]).clip(lower=0)

                    df_bord_view = df_bord[
                        ["bordado_codigo", "bordado_nome", "qtd_necessaria",
                         "estoque_atual", "falta"]
                    ].sort_values("bordado_nome")

                    st.dataframe(df_bord_view)

                    excel_bord = gerar_excel_simples(df_bord_view, sheet_name="Bordados")
                    st.download_button(
                        "ðŸ’¾ Baixar 'Produzir Hoje â€” Bordados' (Excel)",
                        data=excel_bord,
                        file_name="produzir_hoje_bordados.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                # --------------------------------------------------------------
                # 2.5. EXPLICAÃ‡ÃƒO FINAL
                # --------------------------------------------------------------
                st.markdown("---")
                st.markdown(
                    """
                    ### ðŸ§­ Resumo do fluxo

                    1. **Leitura do template_estoque**  
                       - Direto do Google Sheets, aba configurada.  
                       - Somente leitura, nenhuma cÃ©lula Ã© alterada.

                    2. **Leitura da planilha de vendas**  
                       - VocÃª sobe um XLSX diÃ¡rio (Mercado Livre, Shopee, etc.).  
                       - O app soma por cÃ³digo.

                    3. **Produtos Prontos Faltantes**  
                       - Compara vendas x estoque_atual.  
                       - SÃ³ explode em insumos o que realmente estÃ¡ faltando.

                    4. **ExplosÃ£o em Semi / Gola / Bordado**  
                       - Respeita kits (`eh_kit`, `componentes`, `quantidades`).  
                       - Usa `semi_codigo`, `gola_codigo`, `bordado_codigo`
                         para montar o plano de produÃ§Ã£o.

                    5. **RelatÃ³rios de ProduÃ§Ã£o**  
                       - `Produzir Hoje â€” Semis & Golas` â†’ ideal para cÃ©lula de costura.  
                       - `Produzir Hoje â€” Bordados` â†’ ideal para cÃ©lula de bordado/rebater golas.

                    Tudo isso mantendo o **template_estoque intocÃ¡vel**.
                    """
                )

        except Exception as e:
            st.error(f"Ocorreu um erro ao processar as vendas: {e}")
